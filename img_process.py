
import cv2
import numpy as np
from operator import itemgetter
from statistics import multimode, mean, stdev
from openpyxl import load_workbook, Workbook
from openpyxl.chart.trendline import Trendline
import matplotlib.pyplot as plt
from sklearn.metrics import r2_score
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
from openpyxl.chart.trendline import Trendline
from openpyxl import load_workbook
from openpyxl.chart.marker import Marker


def increase_brightness(img, value=30):
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    h, s, v = cv2.split(hsv)

    lim = 255 - value
    h[h > lim] = 255
    h[h <= lim] += value
    s[s > lim] = 255
    s[s <= lim] += value
    v[v > lim] = 255
    v[v <= lim] += value

    final_hsv = cv2.merge((h, s, v))
    img = cv2.cvtColor(final_hsv, cv2.COLOR_HSV2BGR)
    return img


def absFn(a, b):
    if a > b: return (a - b)
    return (b - a)


def process():
    f = open("file_path.txt", "r")
    path = f.read()
    path1 = path.split('\n')
    filename = path1[0].split('.')
    img = cv2.imread(path1[0])

    # ....................................ปรับภาพ............................................

    ret, thresh = cv2.threshold(cv2.cvtColor(img, cv2.COLOR_BGR2GRAY), 127, 255, cv2.THRESH_BINARY)
    contours, hier = cv2.findContours(thresh, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

    crop_img = cv2.resize(img, (400, 600))
    cv2.imwrite(filename[0] + '_detect' + ".jpg", crop_img)

    img2 = cv2.imread(filename[0] + '_detect' + ".jpg")

    img_color = img2.copy()
    img_blur = cv2.GaussianBlur(img_color, (5, 5), 0)
    img_bright = increase_brightness(img_blur, 50)
    img_gray = cv2.cvtColor(img_bright, cv2.COLOR_RGB2GRAY)
    img_hist = cv2.equalizeHist(img_gray)
    img_mblur = cv2.medianBlur(img_hist, 5)
    edges = cv2.Canny(img_mblur, 100, 200)
    circles = cv2.HoughCircles(edges, cv2.HOUGH_GRADIENT, 1, 37, param1=50, param2=20, minRadius=10, maxRadius=20)
    circles = np.uint16(np.around(circles))

    # print(circles)
    # ...........................................OLD....................................................#
    maxCol = 8
    maxRow = 12
    numTest = 4
    circles = np.uint16(np.around(circles))
    cir = np.array(sorted(circles[0], key=itemgetter(1)))
    rows = cir.shape[0]
    cir_new = cir

    v = [0 for x in range(rows)]
    v0 = [0 for x in range(rows)]
    i = 0
    j = 0
    s = 0
    while (i + numTest - 1 < maxCol):
        p1 = cir[i][:]
        ii = i + numTest - 1
        p2 = cir[ii][:]
        if absFn(p2[1], p1[1]) < 20:
            j = j + 1
        else:
            j = 0
        v[i] = j
        v0[i] = cir[i][0]
        i = i + 1

    z = rows - numTest + 1

    while (i < z):
        p1 = cir[i][:]
        ii = i + numTest - 1
        p2 = cir[ii][:]
        iii = i - numTest + 2
        p3 = cir[iii][:]
        if (absFn(p2[1], p1[1]) < 20):
            if j < maxCol:
                j = j + 1
            else:
                j = 1
        elif absFn(p1[1], p3[1]) < 20:
            if j < maxCol:
                j = j + 1
            else:
                j = 1
        else:
            j = 0
        v[i] = j
        v0[i] = cir[i][0]
        i = i + 1

    while (i < rows):
        p1 = cir[i][:]
        ii = i - numTest + 2
        p3 = cir[ii][:]
        if absFn(p1[1], p3[1]) < 20:
            j = j + 1
        else:
            j = 0
        v[i] = j
        v0[i] = cir[i][0]
        i = i + 1

    pos = [i for i, x in enumerate(v) if x > 0]
    cir = cir[pos][:]
    cir2 = np.array(sorted(cir, key=itemgetter(0)))
    cir_new = [[0, 0, 0]]
    v = [x for i, x in enumerate(v) if x > 0]
    pc = np.concatenate(cir[:][:, 0:1])
    pr = np.concatenate(cir[:][:, 1:2])
    minF = np.min(pc)
    minFr = np.min(pr)
    maxL = np.max(pc)
    jj = 0
    for k in range(maxRow):
        v = [0 for i in range(maxCol)]
        pos = k * maxCol - jj
        sortCir1 = np.uint16(np.array(sorted(cir[pos:pos + maxCol, :], key=itemgetter(0))))
        flg = 1
        while (flg == 1):
            sortCir1 = np.uint16(np.array(sorted(sortCir1, key=itemgetter(0))))
            pc = np.concatenate(sortCir1[:][:, 0:1])
            pr = np.concatenate(sortCir1[:][:, 1:2])
            pcDiff = np.diff(pc)
            prDiff = np.diff(pr)
            valMin = sortCir1.min(axis=0)
            valMax = sortCir1.max(axis=0)
            minX = valMin[0]
            tmp = np.max(pr) - np.min(pr)
            p2 = np.array(np.where(pcDiff == np.uint16(np.max(pcDiff))))[0]

            if ((np.max(pcDiff) > 20 * 3.5) or (tmp > 20)):
                p = np.array(np.where(sortCir1 == np.uint16(np.max(pr))))[0]
                v[p[0]] = p[0]
                sortCir1[p[0]][1] = np.mean(pr)
                if (np.max(pcDiff) > 20 * 3.5):
                    sortCir1[p[0]][0] = np.uint16((pc[p2[0]] + pc[p2[0] + 1]) / 2)
                elif (p[0] > 0) and (p[0] < maxCol - 1):
                    sortCir1[p[0]][0] = np.uint16((pc[p[0] - 1] + pc[p[0] + 1]) / 2)
                elif (p[0] == 0):
                    sortCir1[p[0]][0] = np.uint16(minF)
                elif (p[0] == maxCol - 1):
                    sortCir1[p[0]][0] = np.uint16(maxL)
                jj = jj + 1

            else:
                flg = 0

        sortCir1 = np.uint16(np.array(sorted(sortCir1, key=itemgetter(0))))
        cir_new = np.vstack([cir_new, sortCir1])

    cir = cir_new[0:cir_new.shape[0]][:]

    # .......................................Edit....................................#
    set_x = [0 for x in range(12)]
    set_cols = []
    cir_sortx = np.array(sorted(cir, key=itemgetter(0)))

    a = 0
    for i in range(0, 8):
        b = 0
        for j in range(a, a + 12):
            set_x[b] = cir_sortx[j][0]
            b = b + 1
        var = round(mean(multimode(set_x)))
        set_cols.append(var)
        a = a + 12

    set_y = [0 for x in range(8)]
    set_rows = []
    cir_sorty = np.array(sorted(cir, key=itemgetter(1)))

    c = 0
    for i in range(0, 12):
        d = 0
        for j in range(c, c + 8):
            set_y[d] = cir_sorty[j][1]
            d = d + 1
        var = round(mean(multimode(set_y)))
        set_rows.append(var)
        c = c + 8

    radius = []
    for i in cir:
        radius.append(i[2])
    r = round(mean(multimode(radius)))

    array_cir = [[0 for x in range(3)] for y in range(96)]
    var = 0
    for i in range(0, 8):
        for j in range(0, 12):
            array_cir[var][0] = set_cols[i]
            array_cir[var][1] = set_rows[j]
            array_cir[var][2] = r
            var = var + 1

    rows = 96
    columns = 5
    mylist = [[0 for x in range(columns)] for y in range(rows)]
    j = 0
    for i in array_cir:
        mylist[j][0] = i[0]
        mylist[j][1] = i[1]

        mylist[j][2] = (int(img2[i[1], i[0], 0])
                        + int(img2[i[1] - 1, i[0] + 1, 0])
                        + int(img2[i[1], i[0] + 1, 0])
                        + int(img2[i[1] + 1, i[0] + 1, 0])
                        + int(img2[i[1] + 1, i[0], 0])
                        + int(img2[i[1] + 1, i[0] - 1, 0])
                        + int(img2[i[1], i[0] - 1, 0])
                        + int(img2[i[1] - 1, i[0] - 1, 0])
                        + int(img2[i[1] - 1, i[0], 0])) / 9

        mylist[j][3] = (int(img2[i[1], i[0], 1])
                        + int(img2[i[1] - 1, i[0] + 1, 1])
                        + int(img2[i[1], i[0] + 1, 1])
                        + int(img2[i[1] + 1, i[0] + 1, 1])
                        + int(img2[i[1] + 1, i[0], 1])
                        + int(img2[i[1] + 1, i[0] - 1, 1])
                        + int(img2[i[1], i[0] - 1, 1])
                        + int(img2[i[1] - 1, i[0] - 1, 1])
                        + int(img2[i[1] - 1, i[0], 1])) / 9

        mylist[j][4] = (int(img2[i[1], i[0], 2])
                        + int(img2[i[1] - 1, i[0] + 1, 2])
                        + int(img2[i[1], i[0] + 1, 2])
                        + int(img2[i[1] + 1, i[0] + 1, 2])
                        + int(img2[i[1] + 1, i[0], 2])
                        + int(img2[i[1] + 1, i[0] - 1, 2])
                        + int(img2[i[1], i[0] - 1, 2])
                        + int(img2[i[1] - 1, i[0] - 1, 2])
                        + int(img2[i[1] - 1, i[0], 2])) / 9

        # draw the outer circle
        cv2.circle(img2, (i[0], i[1]), i[2], (0, 255, 0), 2)
        # draw the center of the circle
        cv2.circle(img2, (i[0], i[1]), 1, (0, 0, 255), 3)
        j = j + 1

    r = 0;
    jj = 0
    for i in range(1, 97):
        cv2.putText(img2, str(i), (minF + jj * 45, minFr - 20 + r * 45), cv2.FONT_HERSHEY_SIMPLEX, 0.4, 0)
        jj = jj + 1
        if (jj == maxCol):
            jj = 0
            ####number columns####
            cv2.putText(img2, "(" + str(r + 1) + ")", (minF - 10, minFr + r * 45), cv2.FONT_HERSHEY_SIMPLEX, 0.4,
                        (255, 0, 0))
            r = r + 1
    cv2.imwrite(filename[0] + '_detect' + ".jpg", img2)
    # cv2.imshow('img',img2)
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()

    wb = Workbook()
    sheet1 = wb.active

    # lab information
    sheet1.title = "Sheet2"
    with open("file_labInfo.txt", "r+", encoding="utf-8") as f:
        x = 0
        line = f.read()
        data = line.split("\n")
        for j in range(7):
            data1 = sheet1.cell(row=j + 1, column=1)
            data1.value = data[j]

    sheet2 = wb.create_sheet("Sheet_B", 0)
    sheet2.title = "Sheet1"
    sheet2["A1"] = "sequence"
    sheet2["B1"] = "x"
    sheet2["C1"] = "y"
    sheet2["D1"] = "R"
    sheet2["E1"] = "G"
    sheet2["F1"] = "B"

    r = 2
    num = 1
    for i in range(0, 12):
        for j in range(0 + i, 96, 12):
            data1 = sheet2.cell(row=r, column=1)
            data1.value = num
            data2 = sheet2.cell(row=r, column=2)
            data2.value = mylist[j][0]
            data3 = sheet2.cell(row=r, column=3)
            data3.value = mylist[j][1]
            data4 = sheet2.cell(row=r, column=4)
            data4.value = mylist[j][2]
            data5 = sheet2.cell(row=r, column=5)
            data5.value = mylist[j][3]
            data6 = sheet2.cell(row=r, column=6)
            data6.value = mylist[j][4]
            r = r + 1
            num = num + 1



    wb.save(filename[0] + "_detect" + ".xlsx")
    print("Image Detect Process Successful")
    return (filename[0] + "_detect" + ".xlsx")


def mean_sd_rsd(filepath):
    f = open("file_labInfo.txt", "r", encoding="utf-8")
    word1 = f.read()
    word2 = word1.split("\n")
    chem1 = word2[1].split(": ")
    chem2 = word2[2].split(": ")
    chem3 = word2[3].split(": ")
    chem4 = word2[4].split(": ")
    arrchem = [chem1[1],chem2[1],chem3[1],chem4[1]]

    wb = load_workbook(filepath)
    sheet = wb.active

    text = ["R", "G", "B"]
    red = [0 for x in range(3)]
    green = [0 for x in range(3)]
    blue = [0 for x in range(3)]

    # ...........................rgb no1...........................
    a = 4
    b = 3

    sheet.cell(row=2, column=10).value = arrchem[0]
    sheet.cell(row=2, column=12).value = "Mean"
    sheet.cell(row=2, column=13).value = "SD"
    sheet.cell(row=2, column=14).value = "RSD"
    for i in range(0, 3):
        sheet.cell(row=b, column=9).value = 1
        sheet.cell(row=b, column=10).value = 2
        sheet.cell(row=b, column=11).value = 3
        sheet.cell(row=b, column=13).value = text[i]
        a = a + 1
        b = b + 9

    r = 0
    for j in range(2, 10):
        # red
        sheet.cell(row=r + 4, column=9).value = (sheet.cell(row=j, column=4)).value
        red[0] = (sheet.cell(row=j, column=4)).value
        sheet.cell(row=r + 4, column=10).value = (sheet.cell(row=j + 8, column=4)).value
        red[1] = (sheet.cell(row=j + 8, column=4)).value
        sheet.cell(row=r + 4, column=11).value = (sheet.cell(row=j + 16, column=4)).value
        red[2] = (sheet.cell(row=j + 16, column=4)).value

        # green
        sheet.cell(row=r + 13, column=9).value = (sheet.cell(row=j, column=5)).value
        green[0] = (sheet.cell(row=j, column=5)).value
        sheet.cell(row=r + 13, column=10).value = (sheet.cell(row=j + 8, column=5)).value
        green[1] = (sheet.cell(row=j + 8, column=5)).value
        sheet.cell(row=r + 13, column=11).value = (sheet.cell(row=j + 16, column=5)).value
        green[2] = (sheet.cell(row=j + 16, column=5)).value

        # blue
        sheet.cell(row=r + 22, column=9).value = (sheet.cell(row=j, column=6)).value
        blue[0] = (sheet.cell(row=j, column=6)).value
        sheet.cell(row=r + 22, column=10).value = (sheet.cell(row=j + 8, column=6)).value
        blue[1] = (sheet.cell(row=j + 8, column=6)).value
        sheet.cell(row=r + 22, column=11).value = (sheet.cell(row=j + 16, column=6)).value
        blue[2] = (sheet.cell(row=j + 16, column=6)).value

        # mean
        sheet.cell(row=r + 4, column=12).value = mean(red)
        sheet.cell(row=r + 13, column=12).value = mean(green)
        sheet.cell(row=r + 22, column=12).value = mean(blue)

        # sd
        sheet.cell(row=r + 4, column=13).value = round(stdev(red))
        sheet.cell(row=r + 13, column=13).value = round(stdev(green))
        sheet.cell(row=r + 22, column=13).value = round(stdev(blue))

        # rsd
        rsd_r = (sheet.cell(row=r + 4, column=13).value / sheet.cell(row=r + 4, column=12).value) * 100
        rsd_g = (sheet.cell(row=r + 13, column=13).value / sheet.cell(row=r + 13, column=12).value) * 100
        rsd_b = (sheet.cell(row=r + 22, column=13).value / sheet.cell(row=r + 22, column=12).value) * 100
        sheet.cell(row=r + 4, column=14).value = round(rsd_r)
        sheet.cell(row=r + 13, column=14).value = round(rsd_g)
        sheet.cell(row=r + 22, column=14).value = round(rsd_b)

        r = r + 1

    # ...........................rgb no2...........................
    a = 4
    b = 3
    sheet.cell(row=2, column=20).value = arrchem[1]
    sheet.cell(row=2, column=22).value = "Mean"
    sheet.cell(row=2, column=23).value = "SD"
    sheet.cell(row=2, column=24).value = "RSD"
    for i in range(0, 3):
        sheet.cell(row=b, column=19).value = 4
        sheet.cell(row=b, column=20).value = 5
        sheet.cell(row=b, column=21).value = 6
        sheet.cell(row=b, column=23).value = text[i]
        a = a + 1
        b = b + 9

    r = 0
    for j in range(26, 34):
        # red
        sheet.cell(row=r + 4, column=19).value = (sheet.cell(row=j, column=4)).value
        red[0] = (sheet.cell(row=j, column=4)).value
        sheet.cell(row=r + 4, column=20).value = (sheet.cell(row=j + 8, column=4)).value
        red[1] = (sheet.cell(row=j + 8, column=4)).value
        sheet.cell(row=r + 4, column=21).value = (sheet.cell(row=j + 16, column=4)).value
        red[2] = (sheet.cell(row=j + 16, column=4)).value

        # green
        sheet.cell(row=r + 13, column=19).value = (sheet.cell(row=j, column=5)).value
        green[0] = (sheet.cell(row=j, column=5)).value
        sheet.cell(row=r + 13, column=20).value = (sheet.cell(row=j + 8, column=5)).value
        green[1] = (sheet.cell(row=j + 8, column=5)).value
        sheet.cell(row=r + 13, column=21).value = (sheet.cell(row=j + 16, column=5)).value
        green[2] = (sheet.cell(row=j + 16, column=5)).value

        # blue
        sheet.cell(row=r + 22, column=19).value = (sheet.cell(row=j, column=6)).value
        blue[0] = (sheet.cell(row=j, column=6)).value
        sheet.cell(row=r + 22, column=20).value = (sheet.cell(row=j + 8, column=6)).value
        blue[1] = (sheet.cell(row=j + 8, column=6)).value
        sheet.cell(row=r + 22, column=21).value = (sheet.cell(row=j + 16, column=6)).value
        blue[2] = (sheet.cell(row=j + 16, column=6)).value

        # mean
        sheet.cell(row=r + 4, column=22).value = mean(red)
        sheet.cell(row=r + 13, column=22).value = mean(green)
        sheet.cell(row=r + 22, column=22).value = mean(blue)

        # sd
        sheet.cell(row=r + 4, column=23).value = round(stdev(red))
        sheet.cell(row=r + 13, column=23).value = round(stdev(green))
        sheet.cell(row=r + 22, column=23).value = round(stdev(blue))

        # rsd
        rsd_r = (sheet.cell(row=r + 4, column=23).value / sheet.cell(row=r + 4, column=22).value) * 100
        rsd_g = (sheet.cell(row=r + 13, column=23).value / sheet.cell(row=r + 13, column=22).value) * 100
        rsd_b = (sheet.cell(row=r + 22, column=23).value / sheet.cell(row=r + 22, column=22).value) * 100
        sheet.cell(row=r + 4, column=24).value = round(rsd_r)
        sheet.cell(row=r + 13, column=24).value = round(rsd_g)
        sheet.cell(row=r + 22, column=24).value = round(rsd_b)

        r = r + 1

    # ...........................rgb no3...........................
    a = 36
    b = 35
    sheet.cell(row=34, column=10).value = arrchem[2]
    sheet.cell(row=34, column=12).value = "Mean"
    sheet.cell(row=34, column=13).value = "SD"
    sheet.cell(row=34, column=14).value = "RSD"
    for i in range(0, 3):
        sheet.cell(row=b, column=9).value = 7
        sheet.cell(row=b, column=10).value = 8
        sheet.cell(row=b, column=11).value = 9
        sheet.cell(row=b, column=13).value = text[i]
        a = a + 1
        b = b + 9

    r = 32

    for j in range(50, 58):
        # red
        sheet.cell(row=r + 4, column=9).value = (sheet.cell(row=j, column=4)).value
        red[0] = (sheet.cell(row=j, column=4)).value
        sheet.cell(row=r + 4, column=10).value = (sheet.cell(row=j + 8, column=4)).value
        red[1] = (sheet.cell(row=j + 8, column=4)).value
        sheet.cell(row=r + 4, column=11).value = (sheet.cell(row=j + 16, column=4)).value
        red[2] = (sheet.cell(row=j + 16, column=4)).value

        # green
        sheet.cell(row=r + 13, column=9).value = (sheet.cell(row=j, column=5)).value
        green[0] = (sheet.cell(row=j, column=5)).value
        sheet.cell(row=r + 13, column=10).value = (sheet.cell(row=j + 8, column=5)).value
        green[1] = (sheet.cell(row=j + 8, column=5)).value
        sheet.cell(row=r + 13, column=11).value = (sheet.cell(row=j + 16, column=5)).value
        green[2] = (sheet.cell(row=j + 16, column=5)).value

        # blue
        sheet.cell(row=r + 22, column=9).value = (sheet.cell(row=j, column=6)).value
        blue[0] = (sheet.cell(row=j, column=6)).value
        sheet.cell(row=r + 22, column=10).value = (sheet.cell(row=j + 8, column=6)).value
        blue[1] = (sheet.cell(row=j + 8, column=6)).value
        sheet.cell(row=r + 22, column=11).value = (sheet.cell(row=j + 16, column=6)).value
        blue[2] = (sheet.cell(row=j + 16, column=6)).value

        # mean
        sheet.cell(row=r + 4, column=12).value = mean(red)
        sheet.cell(row=r + 13, column=12).value = mean(green)
        sheet.cell(row=r + 22, column=12).value = mean(blue)

        # sd
        sheet.cell(row=r + 4, column=13).value = round(stdev(red))
        sheet.cell(row=r + 13, column=13).value = round(stdev(green))
        sheet.cell(row=r + 22, column=13).value = round(stdev(blue))

        # rsd
        rsd_r = (sheet.cell(row=r + 4, column=13).value / sheet.cell(row=r + 4, column=12).value) * 100
        rsd_g = (sheet.cell(row=r + 13, column=13).value / sheet.cell(row=r + 13, column=12).value) * 100
        rsd_b = (sheet.cell(row=r + 22, column=13).value / sheet.cell(row=r + 22, column=12).value) * 100
        sheet.cell(row=r + 4, column=14).value = round(rsd_r)
        sheet.cell(row=r + 13, column=14).value = round(rsd_g)
        sheet.cell(row=r + 22, column=14).value = round(rsd_b)

        r = r + 1

    # ...........................rgb no4...........................
    a = 36
    b = 35
    sheet.cell(row=34, column=20).value = arrchem[3]
    sheet.cell(row=34, column=22).value = "Mean"
    sheet.cell(row=34, column=23).value = "SD"
    sheet.cell(row=34, column=24).value = "RSD"
    for i in range(0, 3):
        sheet.cell(row=b, column=19).value = 10
        sheet.cell(row=b, column=20).value = 11
        sheet.cell(row=b, column=21).value = 12
        sheet.cell(row=b, column=23).value = text[i]
        a = a + 1
        b = b + 9

    r = 32
    for j in range(74, 82):
        # red
        sheet.cell(row=r + 4, column=19).value = (sheet.cell(row=j, column=4)).value
        red[0] = (sheet.cell(row=j, column=4)).value
        sheet.cell(row=r + 4, column=20).value = (sheet.cell(row=j + 8, column=4)).value
        red[1] = (sheet.cell(row=j + 8, column=4)).value
        sheet.cell(row=r + 4, column=21).value = (sheet.cell(row=j + 16, column=4)).value
        red[2] = (sheet.cell(row=j + 16, column=4)).value

        # green
        sheet.cell(row=r + 13, column=19).value = (sheet.cell(row=j, column=5)).value
        green[0] = (sheet.cell(row=j, column=5)).value
        sheet.cell(row=r + 13, column=20).value = (sheet.cell(row=j + 8, column=5)).value
        green[1] = (sheet.cell(row=j + 8, column=5)).value
        sheet.cell(row=r + 13, column=21).value = (sheet.cell(row=j + 16, column=5)).value
        green[2] = (sheet.cell(row=j + 16, column=5)).value

        # blue
        sheet.cell(row=r + 22, column=19).value = (sheet.cell(row=j, column=6)).value
        blue[0] = (sheet.cell(row=j, column=6)).value
        sheet.cell(row=r + 22, column=20).value = (sheet.cell(row=j + 8, column=6)).value
        blue[1] = (sheet.cell(row=j + 8, column=6)).value
        sheet.cell(row=r + 22, column=21).value = (sheet.cell(row=j + 16, column=6)).value
        blue[2] = (sheet.cell(row=j + 16, column=6)).value

        # mean
        sheet.cell(row=r + 4, column=22).value = mean(red)
        sheet.cell(row=r + 13, column=22).value = mean(green)
        sheet.cell(row=r + 22, column=22).value = mean(blue)

        # sd
        sheet.cell(row=r + 4, column=23).value = round(stdev(red))
        sheet.cell(row=r + 13, column=23).value = round(stdev(green))
        sheet.cell(row=r + 22, column=23).value = round(stdev(blue))

        # rsd
        rsd_r = (sheet.cell(row=r + 4, column=23).value / sheet.cell(row=r + 4, column=22).value) * 100
        rsd_g = (sheet.cell(row=r + 13, column=23).value / sheet.cell(row=r + 13, column=22).value) * 100
        rsd_b = (sheet.cell(row=r + 22, column=23).value / sheet.cell(row=r + 22, column=22).value) * 100
        sheet.cell(row=r + 4, column=24).value = round(rsd_r)
        sheet.cell(row=r + 13, column=24).value = round(rsd_g)
        sheet.cell(row=r + 22, column=24).value = round(rsd_b)

        r = r + 1

    wb.save(filepath)
    print("Find Mean SD RSD Successful")

def plot_graph(filepath,num1,num2,num3,num4,r2):
    wb = load_workbook(filepath)
    sheet = wb.active

    # ..............................group 1............................
    red_y_1 = [0 for x in range(8)]
    green_y_1 = [0 for x in range(8)]
    blue_y_1 = [0 for x in range(8)]
    r = 4
    for i in range(0, 8):
        # red
        a = sheet.cell(row=4, column=12).value
        b = sheet.cell(row=4 + i, column=12).value
        var1 = ((a - b) / a) * 100
        var2 = round(var1, 2)
        sheet.cell(row=r, column=16).value = var2
        red_y_1[i] = var2

        # green
        a = sheet.cell(row=13, column=12).value
        b = sheet.cell(row=13 + i, column=12).value
        var1 = ((a - b) / a) * 100
        var2 = round(var1, 2)
        sheet.cell(row=r + 9, column=16).value = var2
        green_y_1[i] = var2

        # blue
        a = sheet.cell(row=22, column=12).value
        b = sheet.cell(row=22 + i, column=12).value
        var1 = ((a - b) / a) * 100
        var2 = round(var1, 2)
        sheet.cell(row=r + 18, column=16).value = var2
        blue_y_1[i] = var2

        sheet.cell(row=r, column=15).value = num1[i]
        sheet.cell(row=r + 9, column=15).value = num1[i]
        sheet.cell(row=r + 18, column=15).value = num1[i]
        r = r + 1

    # ..............................group 2............................
    r = 4
    red_y_2 = [0 for x in range(8)]
    green_y_2 = [0 for x in range(8)]
    blue_y_2 = [0 for x in range(8)]
    for i in range(0, 8):
        # red
        a = sheet.cell(row=4, column=22).value
        b = sheet.cell(row=4 + i, column=22).value
        var1 = ((a - b) / a) * 100
        var2 = round(var1, 2)
        sheet.cell(row=r, column=26).value = var2
        red_y_2[i] = var2

        # green
        a = sheet.cell(row=13, column=22).value
        b = sheet.cell(row=13 + i, column=22).value
        var1 = ((a - b) / a) * 100
        var2 = round(var1, 2)
        sheet.cell(row=r + 9, column=26).value = var2
        green_y_2[i] = var2

        # blue
        a = sheet.cell(row=22, column=22).value
        b = sheet.cell(row=22 + i, column=22).value
        var1 = ((a - b) / a) * 100
        var2 = round(var1, 2)
        sheet.cell(row=r + 18, column=26).value = var2
        blue_y_2[i] = var2

        sheet.cell(row=r, column=25).value = num2[i]
        sheet.cell(row=r + 9, column=25).value = num2[i]
        sheet.cell(row=r + 18, column=25).value = num2[i]
        r = r + 1

    # ..............................group 3............................
    r = 36
    red_y_3 = [0 for x in range(8)]
    green_y_3 = [0 for x in range(8)]
    blue_y_3 = [0 for x in range(8)]
    for i in range(0, 8):
        # red
        a = sheet.cell(row=36, column=12).value
        b = sheet.cell(row=36 + i, column=12).value
        var1 = ((a - b) / a) * 100
        var2 = round(var1, 2)
        sheet.cell(row=r, column=16).value = var2
        red_y_3[i] = var2

        # green
        a = sheet.cell(row=45, column=12).value
        b = sheet.cell(row=45 + i, column=12).value
        var1 = ((a - b) / a) * 100
        var2 = round(var1, 2)
        sheet.cell(row=r + 9, column=16).value = var2
        green_y_3[i] = var2

        # blue
        a = sheet.cell(row=54, column=12).value
        b = sheet.cell(row=54 + i, column=12).value
        var1 = ((a - b) / a) * 100
        var2 = round(var1, 2)
        sheet.cell(row=r + 18, column=16).value = var2
        blue_y_3[i] = var2

        sheet.cell(row=r, column=15).value = num3[i]
        sheet.cell(row=r + 9, column=15).value = num3[i]
        sheet.cell(row=r + 18, column=15).value = num3[i]
        r = r + 1

    # ..............................group 4............................
    r = 36
    red_y_4 = [0 for x in range(8)]
    green_y_4 = [0 for x in range(8)]
    blue_y_4 = [0 for x in range(8)]
    for i in range(0, 8):
        # red
        a = sheet.cell(row=36, column=22).value
        b = sheet.cell(row=36 + i, column=22).value
        var1 = ((a - b) / a) * 100
        var2 = round(var1, 2)
        sheet.cell(row=r, column=26).value = var2
        red_y_4[i] = var2

        # green
        a = sheet.cell(row=45, column=22).value
        b = sheet.cell(row=45 + i, column=22).value
        var1 = ((a - b) / a) * 100
        var2 = round(var1, 2)
        sheet.cell(row=r + 9, column=26).value = var2
        green_y_4[i] = var2

        # blue
        a = sheet.cell(row=54, column=22).value
        b = sheet.cell(row=54 + i, column=22).value
        var1 = ((a - b) / a) * 100
        var2 = round(var1, 2)
        sheet.cell(row=r + 18, column=26).value = var2
        blue_y_4[i] = var2

        sheet.cell(row=r, column=25).value = num3[i]
        sheet.cell(row=r + 9, column=25).value = num3[i]
        sheet.cell(row=r + 18, column=25).value = num3[i]
        r = r + 1

    # ....................................plot graph 1...................................
    f = open("file_labInfo.txt", "r", encoding="utf-8")
    word1 = f.read()
    word2 = word1.split("\n")
    chem1 = word2[1].split(": ")
    chem2 = word2[2].split(": ")
    chem3 = word2[3].split(": ")
    chem4 = word2[4].split(": ")

    chart = ScatterChart()
    chart.x_axis.title = "Concentration of "+chem1[1]+" (µM)"
    chart.y_axis.title = '% Inhibition'

    # creat chart

    for i in range(16, 17):
        xvalues = Reference(sheet, min_col=15, min_row=4, max_row=11)
        values = Reference(sheet, min_col=i, min_row=3, max_row=11)
        series = Series(values, xvalues, title="R intensity")
        series.marker = Marker('square')
        series.marker.graphicalProperties.solidFill = "C0504D"
        series.marker.graphicalProperties.line.solidFill = "C0504D"
        series.graphicalProperties.line.noFill = True
        chart.series.append(series)
        if (r2[0] == True):
            line_r = chart.series[0]
            line_r.trendline = Trendline()


        xvalues = Reference(sheet, min_col=15, min_row=13, max_row=20)
        values = Reference(sheet, min_col=i, min_row=12, max_row=20)
        series = Series(values, xvalues, title="G intensity")
        series.marker = Marker('triangle')
        series.marker.graphicalProperties.solidFill = "9BBB59"
        series.marker.graphicalProperties.line.solidFill = "9BBB59"
        series.graphicalProperties.line.noFill = True
        chart.series.append(series)
        if (r2[1] == True):
            line_g = chart.series[1]
            line_g.trendline = Trendline()

        xvalues = Reference(sheet, min_col=15, min_row=22, max_row=29)
        values = Reference(sheet, min_col=i, min_row=21, max_row=29)
        series = Series(values, xvalues, title="B intensity")
        series.marker = Marker('diamond')
        series.marker.graphicalProperties.solidFill = "4F81BD"
        series.marker.graphicalProperties.line.solidFill = "4F81BD"
        series.graphicalProperties.line.noFill = True
        chart.series.append(series)
        if (r2[2] == True):
            line_b = chart.series[2]
            line_b.trendline = Trendline()
    sheet.add_chart(chart, "AB7")


    # ....................................plot graph 2...................................
    chart = ScatterChart()
    chart.x_axis.title = "Concentration of "+chem2[1]+" (µM)"
    chart.y_axis.title = '% Inhibition'

    for i in range(26, 27):
        xvalues = Reference(sheet, min_col=25, min_row=4, max_row=11)
        values = Reference(sheet, min_col=i, min_row=3, max_row=11)
        series = Series(values, xvalues, title="R intensity")
        series.marker = Marker('square')
        series.marker.graphicalProperties.solidFill = "C0504D"
        series.marker.graphicalProperties.line.solidFill = "C0504D"
        series.graphicalProperties.line.noFill = True
        chart.series.append(series)
        if (r2[0] == True):
            line_r = chart.series[0]
            line_r.trendline = Trendline()


        xvalues = Reference(sheet, min_col=25, min_row=13, max_row=20)
        values = Reference(sheet, min_col=i, min_row=12, max_row=20)
        series = Series(values, xvalues, title="G intensity")
        series.marker = Marker('triangle')
        series.marker.graphicalProperties.solidFill = "9BBB59"
        series.marker.graphicalProperties.line.solidFill = "9BBB59"
        series.graphicalProperties.line.noFill = True
        chart.series.append(series)
        if (r2[1] == True):
            line_g = chart.series[1]
            line_g.trendline = Trendline()

        xvalues = Reference(sheet, min_col=25, min_row=22, max_row=29)
        values = Reference(sheet, min_col=i, min_row=21, max_row=29)
        series = Series(values, xvalues, title="B intensity")
        series.marker = Marker('diamond')
        series.marker.graphicalProperties.solidFill = "4F81BD"
        series.marker.graphicalProperties.line.solidFill = "4F81BD"
        series.graphicalProperties.line.noFill = True
        chart.series.append(series)
        if (r2[2] == True):
            line_b = chart.series[2]
            line_b.trendline = Trendline()
    sheet.add_chart(chart, "AB24")

    # ....................................plot graph 3...................................
    chart = ScatterChart()
    chart.x_axis.title = "Concentration of "+chem3[1]+" (µM)"
    chart.y_axis.title = '% Inhibition'
    for i in range(16, 17):
        xvalues = Reference(sheet, min_col=15, min_row=36, max_row=43)
        values = Reference(sheet, min_col=i, min_row=35, max_row=43)
        series = Series(values, xvalues, title="R intensity")
        series.marker = Marker('square')
        series.marker.graphicalProperties.solidFill = "C0504D"
        series.marker.graphicalProperties.line.solidFill = "C0504D"
        series.graphicalProperties.line.noFill = True
        chart.series.append(series)
        if (r2[0] == True):
            line_r = chart.series[0]
            line_r.trendline = Trendline()

        xvalues = Reference(sheet, min_col=15, min_row=45, max_row=52)
        values = Reference(sheet, min_col=i, min_row=44, max_row=52)
        series = Series(values, xvalues, title="G intensity")
        series.marker = Marker('triangle')
        series.marker.graphicalProperties.solidFill = "9BBB59"
        series.marker.graphicalProperties.line.solidFill = "9BBB59"
        series.graphicalProperties.line.noFill = True
        chart.series.append(series)
        if (r2[1] == True):
            line_g = chart.series[1]
            line_g.trendline = Trendline()

        xvalues = Reference(sheet, min_col=15, min_row=54, max_row=61)
        values = Reference(sheet, min_col=i, min_row=53, max_row=61)
        series = Series(values, xvalues, title="B intensity")
        series.marker = Marker('diamond')
        series.marker.graphicalProperties.solidFill = "4F81BD"
        series.marker.graphicalProperties.line.solidFill = "4F81BD"
        series.graphicalProperties.line.noFill = True
        chart.series.append(series)
        if (r2[2] == True):
            line_b = chart.series[2]
            line_b.trendline = Trendline()
    sheet.add_chart(chart, "AB41")

    # ....................................plot graph 4...................................
    chart = ScatterChart()
    chart.x_axis.title = "Concentration of "+chem4[1]+" (µM)"
    chart.y_axis.title = '% Inhibition'
    for i in range(26, 27):
        xvalues = Reference(sheet, min_col=25, min_row=36, max_row=43)
        values = Reference(sheet, min_col=i, min_row=35, max_row=43)
        series = Series(values, xvalues, title="R intensity")
        series.marker = Marker('square')
        series.marker.graphicalProperties.solidFill = "C0504D"
        series.marker.graphicalProperties.line.solidFill = "C0504D"
        series.graphicalProperties.line.noFill = True
        chart.series.append(series)
        if (r2[0] == True):
            line_r = chart.series[0]
            line_r.trendline = Trendline()

        xvalues = Reference(sheet, min_col=25, min_row=45, max_row=52)
        values = Reference(sheet, min_col=i, min_row=44, max_row=52)
        series = Series(values, xvalues, title="G intensity")
        series.marker = Marker('triangle')
        series.marker.graphicalProperties.solidFill = "9BBB59"
        series.marker.graphicalProperties.line.solidFill = "9BBB59"
        series.graphicalProperties.line.noFill = True
        chart.series.append(series)
        if (r2[1] == True):
            line_g = chart.series[1]
            line_g.trendline = Trendline()

        xvalues = Reference(sheet, min_col=25, min_row=54, max_row=61)
        values = Reference(sheet, min_col=i, min_row=53, max_row=61)
        series = Series(values, xvalues, title="B intensity")
        series.marker = Marker('diamond')
        series.marker.graphicalProperties.solidFill = "4F81BD"
        series.marker.graphicalProperties.line.solidFill = "4F81BD"
        series.graphicalProperties.line.noFill = True
        chart.series.append(series)
        if (r2[2] == True):
            line_b = chart.series[2]
            line_b.trendline = Trendline()
    sheet.add_chart(chart, "AB58")

    wb.save(filepath)
    print("Plot Graph Successful")

    # ***************************graph pic*******************
    print("...........................")
    graph_pic(num1, red_y_1, green_y_1, blue_y_1, chem1[1], r2, 1)
    graph_pic(num2, red_y_2, green_y_2, blue_y_2, chem2[1], r2, 2)
    graph_pic(num3, red_y_3, green_y_3, blue_y_3, chem3[1], r2, 3)
    graph_pic(num4, red_y_4, green_y_4, blue_y_4, chem4[1], r2, 4)
    print("...........................")




def graph_pic(num1,red_y_1,green_y_1,blue_y_1,chem,r2,n):
    m1 = max(red_y_1)
    m2 = max(blue_y_1)
    m3 = max(blue_y_1)
    m = max(m1,m2,m3)
    plot = plt.figure(n)
    plt.ylabel('% Inhibition')
    plt.xlabel("Concentration of " + chem + " (µM)")
    plt.plot(num1, red_y_1, 'o', color='red')
    plt.plot(num1, green_y_1, 'o', color='green')
    plt.plot(num1, blue_y_1, 'o', color='blue')

    if (r2[0] == True):
        r = r_squared(num1, red_y_1)
        plt.text(0.5, m, '(Red) R-squared = %0.4f' % r)
        slope, intercept = np.polyfit(num1, red_y_1, 1)
        abline_values = [slope * i + intercept for i in num1]
        plt.plot(num1, abline_values, color='black')
    if (r2[1] == True):
        r = r_squared(num1, green_y_1)
        plt.text(0.5, m-10, '(Green) R-squared = %0.4f' % r)
        slope, intercept = np.polyfit(num1, green_y_1, 1)
        abline_values = [slope * i + intercept for i in num1]
        plt.plot(num1, abline_values, color='black')
    if (r2[2] == True):
        r = r_squared(num1, blue_y_1)
        plt.text(0.5, m-20, '(Blue) R-squared = %0.4f' % r)
        slope, intercept = np.polyfit(num1, blue_y_1, 1)
        abline_values = [slope * i + intercept for i in num1]
        plt.plot(num1, abline_values, color='black')
    plt.savefig('figure' + str(n))
    print("Graph Pic Successful "+ str(n))

def r_squared(x, y):
    correlation_matrix = np.corrcoef(x, y)
    correlation_xy = correlation_matrix[0, 1]
    r2 = correlation_xy ** 2
    return r2

def main():
    #path = r'C:\Users\111COM\Desktop\23-04-01.jpeg'
    #exel_path = process(path)
    #mean_sd_rsd(exel_path)
    #plot_graph(exel_path)
    return 0

if __name__ == "__main__":
    main()
